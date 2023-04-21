from openpyxl.reader.excel import load_workbook

def read_data(rowNum, colNum):
        wrkBk = load_workbook("linkedIn_login_data.xlsx")
        sheet = wrkBk.get_sheet_by_name("Sheet1")
        return sheet.cell(row=rowNum, column=colNum).value



